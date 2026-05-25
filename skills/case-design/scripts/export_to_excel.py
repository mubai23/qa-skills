#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""将 Markdown 测试用例表格导出为 Excel 文件。

用法:
  python export_to_excel.py <markdown_file> [output_dir]

示例:
  python export_to_excel.py testcases/登录功能测试用例.md
  python export_to_excel.py testcases/登录功能测试用例.md ./output

说明:
  - 如果 Markdown 中有多个表格，自动选择列数最多的那张（通常是详细用例表）
  - P0/P1/P2/P3 优先级列自动着色
  - 中文字符按 2 倍宽度计算，列宽更准确
"""

import os
import re
import sys
import unicodedata

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("请先安装 openpyxl: pip install openpyxl")
    sys.exit(1)


# 优先级颜色映射（背景色，ARGB 格式）
PRIORITY_COLORS = {
    "P0": "FFFF4444",  # 红色
    "P1": "FFFFA500",  # 橙色
    "P2": "FFFFFF99",  # 黄色
    "P3": "FFD3D3D3",  # 浅灰
}


def str_display_width(text):
    """计算字符串的显示宽度（中文字符算 2，英文算 1）。"""
    width = 0
    for ch in text:
        eaw = unicodedata.east_asian_width(ch)
        width += 2 if eaw in ("W", "F") else 1
    return width


def parse_all_markdown_tables(content):
    """从 Markdown 内容中提取所有表格，每个表格为一个二维列表。"""
    lines = content.split("\n")
    tables = []
    current_table = []
    in_table = False
    header_sep_found = False

    for line in lines:
        stripped = line.strip()
        is_table_row = stripped.startswith("|") and stripped.endswith("|")

        if is_table_row:
            if not in_table:
                # 新表格开始
                in_table = True
                header_sep_found = False
                current_table = []
                headers = [h.strip() for h in stripped.split("|")[1:-1]]
                current_table.append(headers)
            elif not header_sep_found:
                # 检查分隔行
                if re.match(r"^\|[\s\-:|]+\|$", stripped):
                    header_sep_found = True
                else:
                    # 不是分隔行，当作数据行
                    values = [v.strip() for v in stripped.split("|")[1:-1]]
                    current_table.append(values)
            else:
                # 数据行
                values = [v.strip() for v in stripped.split("|")[1:-1]]
                current_table.append(values)
        else:
            if in_table:
                # 表格结束
                if len(current_table) >= 2:  # 至少有表头和一行数据
                    tables.append(current_table)
                current_table = []
                in_table = False
                header_sep_found = False

    # 处理文件末尾的表格
    if in_table and len(current_table) >= 2:
        tables.append(current_table)

    return tables


def select_main_table(tables):
    """选择最合适的表格：优先选列数最多的（通常是详细用例表）。"""
    if not tables:
        return None
    # 按列数降序排列，列数相同时按行数降序
    return sorted(tables, key=lambda t: (len(t[0]), len(t)), reverse=True)[0]


def find_priority_col(headers):
    """找到优先级列的索引（列名包含"优先级"或"Priority"）。"""
    for i, h in enumerate(headers):
        if "优先级" in h or "priority" in h.lower():
            return i
    return -1


def export_to_excel(rows, output_path, sheet_name="测试用例"):
    """将行数据写入 Excel 文件，支持优先级着色。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    if not rows:
        print("未找到表格数据")
        return None

    headers = rows[0]
    priority_col = find_priority_col(headers)

    # 表头样式
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 数据行
    for row_idx, row_data in enumerate(rows[1:], 2):
        priority_val = ""
        if priority_col >= 0 and priority_col < len(row_data):
            priority_val = row_data[priority_col].strip().upper()

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

            # 优先级列着色
            if col_idx - 1 == priority_col and priority_val in PRIORITY_COLORS:
                cell.fill = PatternFill(
                    start_color=PRIORITY_COLORS[priority_val],
                    end_color=PRIORITY_COLORS[priority_val],
                    fill_type="solid"
                )

        # 整行浅色背景（交替行，提升可读性）
        if row_idx % 2 == 0 and not priority_val:
            pass  # 保持白色背景

    # 自适应列宽（中文字符按 2 倍宽度）
    for col_idx in range(1, len(headers) + 1):
        max_width = 8
        for row_data in rows:
            if col_idx <= len(row_data):
                cell_text = str(row_data[col_idx - 1])
                for line in cell_text.split("\n"):
                    max_width = max(max_width, str_display_width(line))
        # 限制最大列宽，避免过宽
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_width + 2, 60)

    # 固定表头行
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 28

    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    wb.save(output_path)
    return output_path


def main():
    if len(sys.argv) < 2:
        print("用法: python export_to_excel.py <markdown_file> [output_dir]")
        sys.exit(1)

    md_file = sys.argv[1]
    if not os.path.exists(md_file):
        print(f"文件不存在: {md_file}")
        sys.exit(1)

    output_dir = sys.argv[2] if len(sys.argv) > 2 else os.path.dirname(os.path.abspath(md_file))
    base_name = os.path.splitext(os.path.basename(md_file))[0]
    output_path = os.path.join(output_dir, f"{base_name}.xlsx")

    with open(md_file, "r", encoding="utf-8") as f:
        content = f.read()

    tables = parse_all_markdown_tables(content)
    if not tables:
        print("未找到有效的 Markdown 表格数据")
        sys.exit(1)

    if len(tables) > 1:
        print(f"发现 {len(tables)} 个表格，自动选择列数最多的表格（通常是详细用例表）")

    main_table = select_main_table(tables)
    if not main_table or len(main_table) < 2:
        print("所选表格数据不足（至少需要表头 + 1 行数据）")
        sys.exit(1)

    result = export_to_excel(main_table, output_path)
    if result:
        print(f"已导出 {len(main_table) - 1} 条用例到: {result}")
        print(f"优先级着色：P0=红 / P1=橙 / P2=黄 / P3=灰")


if __name__ == "__main__":
    main()
